import csv
from openpyxl import load_workbook

def match_genes_and_replace(input_gene_coordinates_file, input_counts_file, output_file):
    gene_coordinates = {}  # Dicionário para armazenar os genes e suas coordenadas

    # Lendo a tabela de genes e coordenadas e armazenando os dados relevantes
    with open(input_gene_coordinates_file, 'r') as gene_coord_file:
        reader = csv.reader(gene_coord_file, delimiter='\t')
        for row in reader:
            gene_name = row[0]
            coordinates = row[1].strip("'")  # Removendo as aspas simples
            gene_coordinates[gene_name] = coordinates

    # Lendo a tabela de counts e substituindo os nomes dos genes pelas coordenadas correspondentes
    wb = load_workbook(filename=input_counts_file)
    ws = wb.active
    with open(output_file, 'w', newline='') as outfile:
        writer = csv.writer(outfile, delimiter='\t')
        for row in ws.iter_rows(values_only=True):
            gene_info = row[0].split()[-1]  # Pegando a parte final da string para encontrar as coordenadas
            for gene, coord in gene_coordinates.items():
                if coord in gene_info:
                    new_gene_name = f"{gene}-{coord}"
                    row_list = list(row)  # Convertendo a tupla em uma lista
                    row_list[0] = new_gene_name
                    writer.writerow(row_list)
                    break


# Exemplo de uso
input_gene_coordinates_file = 'c:\\Users\\Administrator\\Desktop\\Ortogolos\\resultado_anotacao_ortologos.tsv'  # Substitua pelo caminho do arquivo com os nomes de genes e coordenadas
input_counts_file = 'c:\\Users\\Administrator\\Desktop\\ArquivosR\\countdatatemp.xlsx'  # Substitua pelo caminho do arquivo de counts em formato xlsx
output_file = 'c:\\Users\\Administrator\\Desktop\\Ortogolos\\counts_corrigidostemp.tsv'  # Nome do arquivo de saída

match_genes_and_replace(input_gene_coordinates_file, input_counts_file, output_file)
